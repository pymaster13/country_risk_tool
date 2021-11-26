from django import forms

import csv
import io
import openpyxl

from .models import Metric, Category, Country, Link


class UploadFileForm(forms.Form):
    file = forms.FileField()

    def process_data(self):
        file = self.cleaned_data['file']
        if file:
            ext = file.name.split('.')[-1]

            if ext == 'xls' or ext == 'xlsx':
                wb = openpyxl.load_workbook(self.cleaned_data['file'].file, data_only=True)
                data = wb.get_active_sheet()
                excel_data = list()
                criterion = data.cell(1,3)

                for row in data.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    excel_data.append(row_data)

                if(criterion.value == 'Units'):
                    
                    categories = []
                    categories_set = Category.objects.values('name')
                    categories_dict_list = list(categories_set)
                    
                    for cat in categories_dict_list:
                        categories.append(cat['name'])
                    
                    metrics = []
                    metrics_set = Metric.objects.values('code')
                    metrics_dict_list = list(metrics_set)
                    
                    for metric in metrics_dict_list:
                        metrics.append(metric['code'])
                    
                    for metric in excel_data[1:]:
                        if(metric[5] not in categories and metric[5] != 'None'):
                            cat = Category()
                            cat.name = metric[5]
                            categories.append(metric[5])
                            cat.save()
                    
                        if(metric[0] not in metrics):
                            temp = Metric()
                    
                            if(metric[5] != 'None'):
                                temp.category = Category.objects.get(name = metric[5])
                            temp.code = metric[0]
                            temp.name = metric[1]
                            temp.units = metric[2]
                            temp.decription = metric[3]
                            temp.source = metric[4]
                    
                            if metric[6] == 'None' or metric[6] == '':
                                temp.esg = ''
                            else:
                                temp.esg = metric[6]
                    
                            if metric[7] == 'None' or metric[7] == '':
                                temp.sdg = ''
                            else:
                                temp.sdg = metric[7]
                    
                            if metric[8] == 'None' or metric[8] == '':
                                temp.descending = ''
                            else:
                                temp.descending = metric[8]
                            if metric[9] == 'None' or metric[9] == '':
                    
                                temp.default = ''
                            else:
                                temp.default = metric[9]
                            metrics.append(metric[0])
                            temp.save()

                elif(criterion.value == 'GDPPerCapita'):
                    fields = excel_data[0]
                    countries = []
                    countries_set = Country.objects.values('iso')
                    countries_dict_list = list(countries_set)
                    
                    for country in countries_dict_list:
                        countries.append(country['iso'])
                    
                    metrics = []
                    metrics_set = Metric.objects.values('code')
                    metrics_dict_list = list(metrics_set)
                    
                    for metric in metrics_dict_list:
                        metrics.append(metric['code'])
                    
                    for country in excel_data[1:]:
                        if country[0] not in countries:
                            countr = Country()
                            countr.iso = country[0]
                            countr.name = country[1]
                            countries.append(country[0])
                            countr.save()
                        else:
                            countr = Country.objects.get(iso = country[0])
                       
                        for field in fields[2:]:
                            if field not in metrics:
                                metr = Metric()
                                metr.code = field
                                metr.save()
                                metrics.append(field)
                            else:
                                metr = Metric.objects.get(code = field)
                            
                            if (country[fields.index(field)] != '' and country[fields.index(field)] != 'None'):
                                float_metric = float(country[fields.index(field)])
                                try:
                                    link_in_db = Link.objects.get(country = countr, metric = metr, category = metr.category)
                                    if(link_in_db.value != float_metric):
                                        link_in_db.value = float_metric
                                        link_in_db.save()
                                except:
                                    link = Link(country = countr, metric = metr, category = metr.category, value = float_metric)
                                    link.save()

                elif(criterion.value == 'LongBondYield'):
                    fields = excel_data[0]
                    
                    for item in excel_data[1:]:
                        country = Country.objects.get(iso = item[0])
                        category = Category.objects.get(name = 'Valuation')
                        
                        for field in fields[2:]:
                            
                            if (item[fields.index(field)] != '' and item[fields.index(field)] != 'None'):
                                metric = Metric.objects.get(code = field)
                                value = float(item[fields.index(field)])

                                try:
                                    link = Link.objects.get(country = country, metric = metric, category = category)

                                    if(link.value != value):
                                        link.value = value
                                        link.save()
                                    continue

                                except:
                                    link = Link(country = country, metric = metric, category = category, value = value)
                                    link.save()


            elif ext == 'csv':

                    f = io.TextIOWrapper(self.cleaned_data['file'].file)
                    data = csv.DictReader(f)

                    if (data.fieldnames[2] == "Units"):
                        categories = []
                        categories_set = Category.objects.values('name')
                        categories_dict_list = list(categories_set)
                        
                        for cat in categories_dict_list:
                            categories.append(cat['name'])
                        
                        metrics = []
                        metrics_set = Metric.objects.values('code')
                        metrics_dict_list = list(metrics_set)
                        
                        for metric in metrics_dict_list:
                            metrics.append(metric['code'])
                        
                        for metric in data:
                            if(metric['Category'] not in categories and metric['Category'] != ""):
                                cat = Category()
                                cat.name = metric["Category"]
                                categories.append(metric["Category"])
                                cat.save()
                        
                            if(metric['Code'] not in metrics):
                                temp = Metric()
                        
                                if(metric['Category'] != ""):
                                    temp.category = Category.objects.get(name = metric['Category'])
                        
                                temp.code = metric['Code']
                                temp.name = metric['Name']
                                temp.units = metric['Units']
                                temp.source = metric['Source']
                                temp.decription = metric['Decription']
                                temp.esg = metric['ESG']
                                temp.sdg = metric['SDG']
                                temp.descending = metric['Descending']
                                temp.default = metric['Default']
                        
                                metrics.append(metric['Code'])
                        
                                temp.save()


                    elif (data.fieldnames[2] == "GDPPerCapita"):
                        fields = data.fieldnames
                        countries = []
                        countries_set = Country.objects.values('iso')
                        countries_dict_list = list(countries_set)
                        
                        for country in countries_dict_list:
                            countries.append(country['iso'])
                        
                        metrics = []
                        metrics_set = Metric.objects.values('code')
                        metrics_dict_list = list(metrics_set)
                        
                        for metric in metrics_dict_list:
                            metrics.append(metric['code'])
                        
                        for country in data:
                            if country[fields[0]] not in countries:
                                countr = Country()
                                countr.iso = country[fields[0]]
                                countr.name = country[fields[1]]
                                countries.append(country[fields[0]])
                                countr.save()
                            else:
                                countr = Country.objects.get(iso = country[fields[0]])
                        
                            for field in fields[2:]:
                                if field not in metrics:
                                    metr = Metric()
                                    metr.code = field
                                    metr.save()
                                    metrics.append(field)
                                else:
                                    metr = Metric.objects.get(code = field)
                        
                                if country[field] != '':
                                    float_metric = float(country[field])
                        
                                    try:
                                        link_in_db = Link.objects.get(country = countr, metric = metr, category = metr.category)
                                        if(link_in_db.value == float_metric):
                                            pass
                                        else:
                                            link_in_db.value = float_metric
                                            link_in_db.save()
                                    except:
                                        link = Link(country = countr, metric = metr, category = metr.category, value = float_metric)
                                        link.save()


                    elif (data.fieldnames[2] == "LongBondYield"):
                        fields = data.fieldnames
                        
                        for item in data:
                            country = Country.objects.get(iso = item[fields[0]])
                            category = Category.objects.get(name = 'Valuation')
                        
                            for field in fields[2:]:
                                if item[field] != "":
                                    metric = Metric.objects.get(code = field)
                                    value = float(item[field])
                        
                                    try:
                                        link = Link.objects.get(country = country, metric = metric, category = category)
                                        if(link.value == value):
                                            pass
                                        else:
                                            link.value = value
                                            link.save()
                                        continue
                        
                                    except:
                                        link = Link(country = country, metric = metric, category = category, value = value)
                                        link.save()
